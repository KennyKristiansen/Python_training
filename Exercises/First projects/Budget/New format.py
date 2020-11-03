import time
import os
import openpyxl as excel
import pprint
import ast

project_folder = r'C:\Users\kenny\PycharmProjects\Python training\Exercises\First projects\Budget'
local_folder = r'C:\Users\kenny\Documents\Python projects\Budget'
os.chdir(local_folder)


files = os.listdir(local_folder)
file_paths = []

for i in range(len(files)):
    if str(files[i]).endswith('.xlsx'):

        file_paths += [files[i]]
        # print('File found: ' + files[i])
        # noinspection PyBroadException
        try:
            check = os.path.exists(files[i])  # Check if file exists in the right folder.
            # print('File found: ' + files[i])
        except Exception as exc:
            print('failed')


def excel_get_values(it, col, sheet):
    cell_value = sheet.cell(row=it, column=col).value
    if cell_value is not None:
        return cell_value


def excel_lib_decoder():
    os.chdir(local_folder)
    start = time.perf_counter()  # Timetaking
    dictionary = {}

    file = open('budget.txt', 'w', encoding='utf-8')

    # Decoding files
    transaction_number = 1
    for a in range(len(file_paths)):
        konto_name = os.path.splitext(file_paths[a])[0]
        print(konto_name, end=': ')
        wb = excel.load_workbook(file_paths[a])
        sheet = wb[wb.sheetnames[0]]
        rows = sheet.max_row - 1
        print('Found {0} transactions.'.format(str(rows)))
        information = {}

        for ii in range(1, rows):
            date = excel_get_values(ii, 1, sheet)  # Dato
            transaction = excel_get_values(ii, 3, sheet)  # Overførsel til
            cash_moved = excel_get_values(ii, 4, sheet)  # Beløb flyttet
            cash_left = excel_get_values(ii, 5, sheet)  # Resterende beløb

            information[transaction_number] = {'Date': str(date), 'Transaction': str(transaction),
                                               'Cash moved': str(cash_moved), 'Cash left': str(cash_left)}
            transaction_number += 1
        dictionary[konto_name] = information
    file.write(str(dictionary))
    file.close()
    stop = time.perf_counter()
    print('Time to decode: ' + str(round(stop - start, 2)) + ' seconds')


def budget_collected():
    os.chdir(local_folder)
    budget_file = open('budget.txt', 'r', encoding='utf-8')
    budget_lines = budget_file.readlines()
    dictionary = dict(ast.literal_eval(budget_lines[0]))
    # print(dictionary.keys())
    budget_file.close()
    transactions = 0
    transactions_id = []
    # Find all transactions IDs for summing transactions to same place
    for accounts in range(len(dictionary)):
        account = list(dictionary.keys())[accounts]
        # pprint.pprint(dictionary[account])
        transactions += len(dictionary[account])
        for ids in range(len(dictionary[account])):
            print(ids)
            print(dictionary[account][ids+1])
            transactions_id.append(str(list(dictionary[account])[ids]))  # This returns all id numbers
            # print(list(dictionary[account].values())[ids])

    # print(transactions_id)


excel_lib_decoder()
budget_collected()


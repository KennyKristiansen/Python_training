import os
import logging
# import datetime
import ast
import pprint
import time
from Budget_GUI import Ui_MainWindow  # importing our generated file
import sys
from PyQt5 import QtGui, QtWidgets
from functools import partial

# Setting work folder
project_folder = r'C:\Users\kenny\PycharmProjects\Python training\Exercises\First projects\Budget'
local_folder = r'C:\Users\kenny\Documents\Python projects\Budget'

categorized = {'bolig': [], 'mad': [], 'transport': [], 'oevrige': [], 'diverse': [], 'opsparing': [], 'gaeld': [],
               'ukategoriseret': []}
os.chdir(project_folder)
categories_file = open('categories.txt', 'r', encoding='utf-8')
categories_file_lines = categories_file.readlines()
categories_file.close()
# noinspection PyBroadException
try:
    categories = dict(ast.literal_eval(categories_file_lines[0]))
except Exception as exc:
    logging.exception('File corrupted, writing new.')
    categories = {'bolig': ['bolig', 'husleje'], 'mad': ['FOETEX'], 'transport': ['dronningemaens'], 'oevrige': [],
                  'diverse': ['avallax', 'byggecenter'], 'opsparing': [], 'gaeld': [], 'ukategoriseret': []}
    categories_file = open('categories.txt', 'w', encoding='utf-8')
    categories_file.write(str(categories))

os.chdir(local_folder)
# Finding files
files = os.listdir(local_folder)
file_paths = []
print('Looking for eligible files.')
for i in range(len(files)):
    if str(files[i]).endswith('.xlsx'):

        file_paths += [files[i]]
        # print('File found: ' + files[i])
        # noinspection PyBroadException
        try:
            check = os.path.exists(files[i])  # Check if file exists in the right folder.
            print('File found: ' + files[i])
        except Exception as exc:
            logging.exception('Path not found: ' + os.getcwd() + str(file_paths[0]))


def excel_get_values(it, col, sheet):
    cell_value = sheet.cell(row=it, column=col).value
    if cell_value is not None:
        return cell_value


def excel_lib_decoder():
    start = time.perf_counter()

    import openpyxl as excel
    # Decoding files
    for a in range(len(file_paths)):
        konto_name = os.path.splitext(file_paths[a])[0]
        # print(konto_name,end= ': ')
        filename = konto_name + '.txt'
        file = open(filename, 'w', encoding='utf-8')
        id = 1
        wb = excel.load_workbook(file_paths[a])
        sheet = wb[wb.sheetnames[0]]
        rows = sheet.max_row - 1
        print('Found {0} transactions.'.format(str(rows)))
        for ii in range(1, rows):
            information = {}
            date = excel_get_values(ii, 1, sheet)  # Dato
            transaction = excel_get_values(ii, 3, sheet)  # Overførsel til
            cash_moved = excel_get_values(ii, 4, sheet)  # Beløb flyttet
            cash_left = excel_get_values(ii, 5, sheet)  # Resterende beløb
            information[id] = {'Date': str(date), 'Transaction': str(transaction), 'Cash moved': str(cash_moved),
                               'Cash left': str(cash_left)}
            id += 1
            file.write((str(information) + '\n'))
        file.close()
    stop = time.perf_counter()
    print('Time to decode: ' + str(round(stop - start,2)) + ' seconds')


def excel_lib_decoder_new():
    start = time.perf_counter()
    complete_dict = {}
    import openpyxl as excel
    file = open('budget.txt', 'w', encoding='utf-8')
    # Decoding files
    for a in range(len(file_paths)):
        konto_name = os.path.splitext(file_paths[a])[0]
        print(konto_name,end= ': ')
        id = 1
        wb = excel.load_workbook(file_paths[a])
        sheet = wb[wb.sheetnames[0]]
        rows = sheet.max_row - 1
        print('Found {0} transactions.'.format(str(rows)))
        information = {}
        for ii in range(1, rows):
            date = excel_get_values(ii, 1, sheet)  # Dato
            transaction = excel_get_values(ii, 3, sheet)  # Overførsel til
            cash_moved = excel_get_values(ii, 4, sheet)  # Beløb flyttet
            cash_left = excel_get_values(ii, 5, sheet)  # Resterende beløb

            information[id] = {'Date': str(date), 'Transaction': str(transaction), 'Cash moved': str(cash_moved),
                               'Cash left': str(cash_left)}
            id += 1
        complete_dict[konto_name] = information
    file.write(str(pprint.pformat(complete_dict)))
    file.close()

    stop = time.perf_counter()
    print('Time to decode: ' + str(round(stop - start,2)) + ' seconds')


def txt_reader():
    global categories_file
    txt_reader_start = time.perf_counter()
    os.chdir(local_folder)
    for iii in range(len(files)):
        categorized_per_file = {'bolig': [], 'mad': [], 'transport': [], 'oevrige': [], 'diverse': [], 'opsparing': [],
                                'gaeld': [], 'ukategoriseret': []}
        if str(files[iii]).endswith('categorized.txt'):
            continue
        if str(files[iii]).endswith('.txt') and files[iii] is not 'geckodriver.txt':
            category_file = False
            file = open(files[iii], 'r', encoding='utf-8')
            konto_name = os.path.splitext(files[iii])[0]
            if not konto_name.endswith('categorized'):
                categories_file = open(konto_name + '_categorized.txt', 'w', encoding='utf-8')
                category_file = True
            else:
                # print('Opening file: ' + files[iii])
                pass
            lines = file.readlines()

            for a in range(0, len(lines)):
                key = ast.literal_eval(lines[a])
                values = list(key.values())[0]
                if category_file:
                    transaction = str(values.get('Transaction')).lower()
                    for ib in range(0, len(categories)):
                        file_found = False
                        list_value = (list(categories.values())[ib])
                        list_key = (list(categories.keys())[ib])

                        for ba in range(len(list_value)):
                            if str(list_value[ba]).lower() in transaction:
                                file_found = True

                        if file_found is not False:

                            # print(str(list_value[ba]) + ' was found and added to ' + list_key)
                            categorized.get(list_key).append(str(values))
                            categorized_per_file.get(list_key).append(str(values))

                            break
                        elif list_key == 'ukategoriseret':
                            categorized.get(list_key).append(str(values))
                            categorized_per_file.get(list_key).append(str(values))

                categories_file.write(str(categorized_per_file))
            # print('Closing file: ' + files[iii])
            file.close()
            if category_file:
                categories_file.close()
    txt_reader_stop = time.perf_counter()
    print('Time to read categorize: ' + str(round(txt_reader_stop - txt_reader_start, 3)) + ' seconds')


##########

def txt_optimizer():
    txt_reader_start = time.perf_counter()

    for iii in range(len(files)):
        os.chdir(local_folder)
        categorized_per_file_optimized = {'bolig': [], 'mad': [], 'transport': [], 'oevrige': [], 'diverse': [],
                                          'opsparing': [],
                                          'gaeld': [], 'ukategoriseret': []}
        if str(files[iii]).endswith('categorized.txt'):
            continue
        if str(files[iii]).endswith('.txt') and files[iii] is not 'geckodriver.txt':
            category_file = False
            file = open(files[iii], 'r', encoding='utf-8')
            konto_name = os.path.splitext(files[iii])[0]
            if not konto_name.endswith('categorized'):
                os.chdir(local_folder)
                os.chdir(".\Optimized")
                categories_file = open(konto_name + '_optimized.txt', 'w', encoding='utf-8')
                category_file = True
            else:
                # print('Opening file: ' + files[iii])
                pass
            lines = file.readlines()
            matching = {}
            for aba in range(0, len(lines)):
                key = ast.literal_eval(lines[aba])
                # print(key)
                values = list(key.values())[0]
                # print(values)
                trans = list(values.values())[1].lower()
                # print(trans + ' ' + str(aba))
                matching[str(aba)] = trans

            for ia in range(len(matching)):
                trans_name = (matching.get(str(ia)))
                for iai in range(len(matching)):
                    trans_comp = matching.get(str(iai))
                    if trans_name == trans_comp:
                        if str(iai) in matching:
                            del matching[str(iai)]

            for a in range(0, len(lines)):
                key = ast.literal_eval(lines[a])
                values = list(key.values())[0]
                if category_file:
                    transaction = str(values.get('Transaction')).lower()
                    for ib in range(0, len(categories)):
                        file_found = False
                        list_value = (list(categories.values())[ib])
                        list_key = (list(categories.keys())[ib])

                        for ba in range(len(list_value)):
                            if str(list_value[ba]).lower() in transaction:
                                file_found = True

                        if file_found is not False:
                            # print(str(list_value[ba]) + ' was found and added to ' + list_key)
                            categorized.get(list_key).append(str(values))
                            categorized_per_file_optimized.get(list_key).append(str(values))
                            break
                        elif list_key == 'ukategoriseret':
                            categorized.get(list_key).append(str(values))
                            categorized_per_file_optimized.get(list_key).append(str(values))

                categories_file.write(str(categorized_per_file_optimized))
            # print('Closing file: ' + files[iii])
            file.close()
            if category_file:
                categories_file.close()
    txt_reader_stop = time.perf_counter()
    print('Time to optimize categories: ' + str(round(txt_reader_stop - txt_reader_start, 3)) + ' seconds')


###########


def write_to_category(item, item_category):
    print(str(item))
    print(str(item).isalnum())

    if str(item).isalnum() == False:  ####################### virker ikke
        return  ###################################
    os.chdir(r'C:\Users\kenny\.PyCharm2018.3\config\scratches\Egne programmer')
    # category = categories.get(item_category)
    # First, find if item is already there.
    file = open('categories.txt', 'w', encoding='utf-8')
    presence = True
    print('Trying')
    for it in range(len(categories)):
        print(it, end='\n')
        print(len(categories) - 1)
        print('')
        if it == (len(categories) - 1):
            presence = False
        for b in range(0, (len(list(categories.values())[it]))):
            print(list(categories.values())[it][b])
            print('HERE')
            if item == list(categories.values())[it][b]:
                presence = False
                break

    print(presence)
    if presence is False:
        print(5)
        categories.get(item_category).append(item)
        print('appending: "' + item + '" to category: ' + item_category)
    file.write(str(categories))
    file.close()


os.chdir(local_folder)


class mywindow(QtWidgets.QMainWindow):
    def button_clicked(self, value):
        text_return = str(self.ui.lineEdit.text())
        print(text_return)
        write_to_category(text_return, value)
        print('Button pressed: ' + str(value) + ', Value saved: ' + str(text_return))
        txt_reader()
        self.ui.lineEdit.setText("")

    def __init__(self):

        super(mywindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        tables = self.findChildren(QtWidgets.QTableWidget)

        button_opsparing = self.ui.button_opsparing
        button_mad = self.ui.button_mad
        button_transport = self.ui.button_transport
        button_faste = self.ui.button_faste
        button_gaeld = self.ui.button_gaeld
        button_bolig = self.ui.button_bolig
        button_diverse = self.ui.button_diverse

        # Textfield data

        button_diverse.clicked.connect(partial(self.button_clicked, 'diverse'))
        button_opsparing.clicked.connect(partial(self.button_clicked, 'opsparing'))
        button_mad.clicked.connect(partial(self.button_clicked, 'mad'))
        button_transport.clicked.connect(partial(self.button_clicked, 'tranport'))
        button_faste.clicked.connect(partial(self.button_clicked, 'oevrige'))
        button_gaeld.clicked.connect(partial(self.button_clicked, 'gaeld'))
        button_bolig.clicked.connect(partial(self.button_clicked, 'bolig'))

        # button_diverse.released.connect(self.ui.lineEdit.clear)
        # button_opsparing.released.connect(self.ui.lineEdit.clear)
        # button_mad.released.connect(self.ui.lineEdit.clear)
        # button_transport.released.connect(self.ui.lineEdit.clear)
        # button_faste.released.connect(self.ui.lineEdit.clear)
        # button_gaeld.released.connect(self.ui.lineEdit.clear)
        # button_bolig.released.connect(self.ui.lineEdit.clear)

        rows = 10
        columns = 2

        ## Get categories
        txt_reader()
        txt_optimizer()
        ### Sort values from category and add them
        count = 0
        dict_col = {}
        # for titles in range(len(categorized.keys())):
        #     title = list(categorized.keys())[titles]
        #     transaction_count = len(categorized.get(title))
        #
        #     for transactions in range(transaction_count):
        #         presence = False
        #         cat = ast.literal_eval(categorized.get(title)[transactions])
        #         #print(categorized.get(title)[transactions])
        #
        #         for transactions_compare in range(transaction_count):
        #             transaction = (cat.get('Transaction'))
        #
        #             #Se om opslagsværdien er den originale værdi
        #             cat_compare = ast.literal_eval(categorized.get(title)[transactions_compare])
        #             transaction_compare = (cat_compare.get('Transaction'))
        #             if str(transaction) == str(transaction_compare):
        #                 date = (cat.get('Date'))
        #                 date_compare = (cat_compare.get('Date'))
        #                 if date != date_compare:
        #                     cash_moved_compare = (cat_compare.get('Cash moved'))
        #                     cash_moved = (cat.get('Cash moved'))
        #                     if cash_moved != cash_moved_compare:
        #                         cash = cash_moved + cash_moved_compare
        #                         cat['Cash moved'] = cash
        #                         cat_compare.popitem()
        #
        #                         #print('found match: ' + str(date) + ' and ' + str(date_compare))
        #                         presence = True

        for i in range(len(tables)):
            table = tables[i]
            table.setRowCount(rows)
            table.setColumnCount(columns)

            table.setFont(QtGui.QFont('Arial', 8))
            table.horizontalHeader().setDefaultSectionSize(150)
            table.horizontalHeader().setMinimumSectionSize(0)
            table.horizontalHeader().setStretchLastSection(True)
            table.horizontalHeader().setSortIndicatorShown(False)

            table.verticalHeader().setDefaultSectionSize(38)
            table.verticalHeader().setMinimumSectionSize(0)
            table.verticalHeader().setStretchLastSection(True)
            table.verticalHeader().setSortIndicatorShown(False)

            table_name = str(table.objectName()).split('_')[0]

            category = categorized.get(table_name)

            top_list_value = []
            top_list_name = []
            top_list_dict = {}

            for aa in range(rows):
                top_list_value.append(0)
                top_list_name.append(0)
            for a in range(len(category)):
                dictionary = ast.literal_eval(str(category[a]))
                value = dictionary.get('Cash moved')
                name = dictionary.get('Transaction')

                for aa in range(rows):
                    if float(value) < float(top_list_value[aa]):
                        top_list_value[aa] = float(value)
                        top_list_name[aa] = name
                        break

            # for a in range(len(category)):
            #     dictionary = ast.literal_eval(category[a])
            #     value = dictionary.get('Cash moved')
            #     name = str(dictionary.get('Transaction'))
            #     for aa in range(rows):
            #         if float(value) < float(top_list_value[aa]):
            #             top_list_value[aa] = float(value)
            #             top_list_name[aa] = name
            #             break

            ##Assign to dictionary

            if len(top_list_value) != 0:
                sorted_values = sorted(zip(top_list_value, top_list_name))
                top_list_value, top_list_name = map(list, zip(*sorted_values))

                for ab in range(len(top_list_value)):
                    # print(ab)
                    # print(top_list_value[ab])
                    # print(top_list_name)

                    top_list_dict[top_list_name[ab]] = top_list_value[ab]

            for x in range(columns):

                for y in range(rows):

                    item = QtWidgets.QTableWidgetItem()
                    if y >= len(list(top_list_dict.values())):
                        break
                    if x == 0:
                        text = str((list(top_list_dict.keys())[y]))  # Beskrivelse
                        item.setText(text)
                    if x == 1:
                        text = str((list(top_list_dict.values())[y]))  # Kroner
                        item.setText(text)
                    table.setItem(y, x, item)


excel_lib_decoder_new()

app = QtWidgets.QApplication([])
# application = mywindow()
# application.showMaximized()
# sys.exit(app.exec())

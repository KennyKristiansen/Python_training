#! python3
# Seamate_converter.py - Converts Seamate data, to excel data

import os
import csv
import re
import configparser
import datetime
import openpyxl as excel
from openpyxl import styles
from openpyxl.styles import Border, Side

program_folder = os.getcwd()
try:
    os.makedirs('Seamate files')
except FileExistsError:
    pass

try:
    os.makedirs('Excel files')
except FileExistsError:
    pass
os.chdir(program_folder)
seamate_folder = os.path.abspath('Seamate files')
excel_folder = os.path.abspath('Excel files')

filenames_in_location = os.listdir(seamate_folder)

matching_files = []
for i in range(len(filenames_in_location)):
    if str(filenames_in_location[i]).startswith('SEAMATE'):
        matching_files += [filenames_in_location[i]]
        # print('Seamate file found: ' + str(filenames_in_location[i]))

# Find newest file, and use that one for further usage
date_extraction = re.compile(r'(SEAMATE_SAMPLE_TESTS_DATA_)(\d\d)(\w)(\d\d)(\w)(\d\d\d\d)(.*)', re.VERBOSE)

newest_file_index = 0
latest_year = 0
latest_month = 0
latest_date = 0

for i in range(len(matching_files)):
    extracted_data = (date_extraction.search(matching_files[i]))
    year = int(extracted_data.group(6))
    if year >= latest_year:
        latest_year = year
        month = int(extracted_data.group(4))
        if month >= latest_month:
            latest_month = month
            date = int(extracted_data.group(2))
            if date >= latest_date:
                latest_date = date

                newest_file_index = i

os.chdir(seamate_folder)
file_open = open(matching_files[newest_file_index])
print('Using file with newest date: ' + str(matching_files[newest_file_index]))
print('Date of file: ' + str(latest_date) + '/' + str(latest_month) + '/' + str(latest_year))
file_read = csv.reader(file_open)

print('Converting Seamate file to Excel')
# Write a category sorted dictionary
dictionary = {}
for row in file_read:
    dictionary_values = {}
    if file_read.line_num > 4:
        # Defining categories
        if file_read.line_num == 5:  # Categories
            categories = row
            # print('Categories are: ' + str(row))
            continue

        if file_read.line_num > 6:

            try:
                bottle_id = row[4]
            except:
                # print('No bottle ID found in row #' + str(file_read.line_num))
                continue
            for values in range(len(row)):
                dictionary_values[categories[values]] = str(row[values])

                dictionary[bottle_id] = dictionary_values

dictionary_keys = list(dictionary.keys())

# save to excel sheet
wb = excel.Workbook(False)
sheet_name = wb.sheetnames[0]
sheet = wb.worksheets[0]

for row_1 in range(len(categories)):
    sheet.cell(row=1, column=((row_1) + 1)).value = categories[row_1]
print('Bottle entries found: ' + str(len(dictionary_keys)))
for row in range(len(dictionary_keys)):
    # print(dictionary.get(dictionary_keys[row]))
    # print(dictionary_keys[row])
    for column in range(len(categories)):
        id_for_bottle = dictionary_keys[row]
        try:
            sheet.cell(row=(row + 2), column=(column + 1)).value = str(dictionary[id_for_bottle][categories[column]])
        except:
            # print('Error happened with id: ' + bottle_id)
            continue
os.chdir(excel_folder)
wb.save('excel_file_name_missing.xlsx')
print('Seamate file succesfully converted to Excel')
wb.close()

os.chdir(excel_folder)
wb = excel.load_workbook('excel_file_name_missing.xlsx', False)
sheet_name = wb.sheetnames[0]
os.chdir(program_folder)
print('Reading settings.txt')
# Read settings file
try:
    settings_file_read = open('settings.txt', mode='r')
except FileNotFoundError:
    print('The settings.txt file was not found.\n\nAn empty settings file has been created.')
    settings_file_write = open('settings.txt', mode='w')
    settings_file_write.write("""
# Instructions:
# This file contains all the adjustable limits for the colorization in the excel spreadsheet.
# If these are left blank, fallback values will be used instead.
# Please do not remove this file, it is needed for running the program.

[Variables]
fe_red =300
fe_orange = 280
fe_yellow=150

v_red =1200
v_orange =1100
v_yellow =1000

s_red =8
s_orange =7
s_yellow =6

ni_red =280
ni_orange =260
ni_yellow =240

cr_red =40
cr_orange =35
cr_yellow =25

pb_red =9
pb_orange =7
pb_yellow =6

cu_red =30
cu_orange =20
cu_yellow =15

ca_red =80000
ca_orange =50000
ca_yellow =40000

zn_red =280
zn_orange =260
zn_yellow =250
""")

config = configparser.ConfigParser()
config.read("settings.txt")

fe_red = config.get("Variables", "fe_red", fallback=300)
fe_orange = config.get("Variables", "fe_orange", fallback=280)
fe_yellow = config.get("Variables", "fe_yellow", fallback=150)
v_red = config.get("Variables", "v_red", fallback=1200)
v_orange = config.get("Variables", "v_orange", fallback=1100)
v_yellow = config.get("Variables", "v_yellow", fallback=1000)
s_red = config.get("Variables", "s_red", fallback=8)
s_orange = config.get("Variables", "s_orange", fallback=7)
s_yellow = config.get("Variables", "s_yellow", fallback=6)
ni_red = config.get("Variables", "ni_red", fallback=280)
ni_orange = config.get("Variables", "ni_orange", fallback=260)
ni_yellow = config.get("Variables", "ni_yellow", fallback=240)
cr_red = config.get("Variables", "cr_red", fallback=40)
cr_orange = config.get("Variables", "cr_orange", fallback=35)
cr_yellow = config.get("Variables", "cr_yellow", fallback=25)
pb_red = config.get("Variables", "pb_red", fallback=9)
pb_orange = config.get("Variables", "pb_orange", fallback=7)
pb_yellow = config.get("Variables", "pb_yellow", fallback=6)
cu_red = config.get("Variables", "cu_red", fallback=30)
cu_orange = config.get("Variables", "cu_orange", fallback=20)
cu_yellow = config.get("Variables", "cu_yellow", fallback=15)
ca_red = config.get("Variables", "ca_red", fallback=80000)
ca_orange = config.get("Variables", "ca_orange", fallback=50000)
ca_yellow = config.get("Variables", "ca_yellow", fallback=40000)
zn_red = config.get("Variables", "zn_red", fallback=280)
zn_orange = config.get("Variables", "zn_orange", fallback=260)
zn_yellow = config.get("Variables", "zn_yellow", fallback=250)

os.chdir(excel_folder)
try:
    wb.sheetnames[1]
except IndexError:
    wb.create_sheet('Monthly', 1)

sheet_data = wb.worksheets[0]
sheet_report = wb.worksheets[1]

print('Identifying latest sample date')
categories = []
for column_count in range(sheet_data.max_column):
    categories.append(str(sheet_data.cell(row=1, column=column_count + 1).value))
group_row = categories.index('GROUP') + 1
date_row = categories.index('SAMPLE DATE') + 1
bottle_row = categories.index('BOTTLE') + 1
location_row = categories.index('LOCATION') + 1
s_row = categories.index('S +/-') + 1
v_row = categories.index('V') + 1
fe_row = categories.index('FE') + 1
ni_row = categories.index('NI') + 1
cr_row = categories.index('CR') + 1
pb_row = categories.index('PB') + 1
cu_row = categories.index('CU') + 1
ca_row = categories.index('CA') + 1
zn_row = categories.index('ZN') + 1

# find latest date

newest_sample = datetime.datetime(1, 1, 1)
latest_sample_date = ()
latest_sample_id = ()
newest_samples = []
latest_date_split = []
for row_count in range(2, sheet_data.max_row):
    date = sheet_data.cell(row=row_count + 1, column=date_row).value
    bottle_id = sheet_data.cell(row=row_count + 1, column=bottle_row).value
    date_split = str.split(date, ' ')

    day = date_split[0]
    if day == 'SAMPLE':
        continue
    month = date_split[1]
    if month == 'Jan':
        month = 1
    if month == 'Feb':
        month = 2
    if month == 'Mar':
        month = 3
    if month == 'Apr':
        month = 4
    if month == 'May':
        month = 5
    if month == 'Jun':
        month = 6
    if month == 'Jul':
        month = 7
    if month == 'Aug':
        month = 8
    if month == 'Sep':
        month = 9
    if month == 'Oct':
        month = 10
    if month == 'Nov':
        month = 11
    if month == 'Dec':
        month = 12

    month = month
    year = date_split[2]
    time = date_split[3]
    # print(day, month, year, time)
    sample_date = datetime.datetime(int(year), int(month), int(day))

    if sample_date > newest_sample:
        newest_samples = []  # Eftersom at nyeste dato er ændret, nulstilles denne liste
        newest_sample = sample_date
        latest_sample_date = sample_date  # Bruges som global værdi for dato
        latest_sample_id = bottle_id  # Bruges som global værdi for ID

    if sample_date == newest_sample:
        newest_samples.append(bottle_id)  # Samler alle de nyeste ID'er i en liste
        latest_date_split = str(str(latest_sample_date).replace(' 00:00:00', '')).split('-')
print('Latest sample date: ' + latest_date_split[2] + '/' + latest_date_split[1] + '/' + latest_date_split[0])

# Excel styles
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))
thick_border = Border(left=Side(style='medium'),
                      right=Side(style='medium'),
                      top=Side(style='medium'),
                      bottom=Side(style='medium'))

red = excel.styles.PatternFill("solid", fgColor="FF0000")
orange = excel.styles.PatternFill("solid", fgColor="FF8000")
yellow = excel.styles.PatternFill("solid", fgColor="FFFF00")
green = excel.styles.PatternFill("solid", fgColor="00FF00")

header_font = excel.styles.Font(bold=True, )
header_alignment = excel.styles.Alignment(horizontal='center')

# Looking up all the wanted values for the newest samples
sheet_report.column_dimensions['A'].width = 8
row = 1
columns_titles = ['location', 'fe', 'v', 's', 'ni', 'cr', 'pb', 'cu', 'ca', 'zn']
for category_count in range(len(columns_titles)):
    sheet_report.cell(row=row, column=category_count + 1).value = columns_titles[category_count]
    sheet_report.cell(row=row, column=category_count + 1).border = thick_border
    sheet_report.cell(row=row, column=category_count + 1).font = header_font
    sheet_report.cell(row=row, column=category_count + 1).alignment = header_alignment

#    sheet_report.cell(row=row, column=(category_count + 1)).value = categories[category_count]

row += 1
for row_count in range(sheet_data.max_row):
    bottle_id = sheet_data.cell(row=row_count + 1, column=bottle_row).value
    bottle_id_correct = False
    for bottle_id_check in range(len(newest_samples)):
        if bottle_id == newest_samples[bottle_id_check]:
            bottle_id_correct = True
    if not bottle_id_correct:
        continue
    location = sheet_data.cell(row=row_count + 1, column=location_row).value
    s = sheet_data.cell(row=row_count + 1, column=s_row).value
    v = sheet_data.cell(row=row_count + 1, column=v_row).value
    fe = sheet_data.cell(row=row_count + 1, column=fe_row).value
    ni = sheet_data.cell(row=row_count + 1, column=ni_row).value
    cr = sheet_data.cell(row=row_count + 1, column=cr_row).value
    pb = sheet_data.cell(row=row_count + 1, column=pb_row).value
    cu = sheet_data.cell(row=row_count + 1, column=cu_row).value
    ca = sheet_data.cell(row=row_count + 1, column=ca_row).value
    zn = sheet_data.cell(row=row_count + 1, column=zn_row).value

    for columns in range(len(columns_titles)):
        column_data = columns_titles[columns]
        # sheet_report.cell(row=int(row), column=int(columns)+1).value = vars()[columns_titles[columns]]

        if columns_titles[columns] == 'location':
            if str(vars()[column_data]) == 'Cylinder Oil Daily Tank':
                sheet_report.cell(row=int(row), column=int(columns) + 1).value = 'Cyl. day tank'
                sheet_report.cell(row=int(row), column=int(columns) + 1).border = thick_border
                sheet_report.cell(row=int(row), column=int(columns) + 1).font = header_font
            else:
                sheet_report.cell(row=int(row), column=int(columns) + 1).value = str(vars()[column_data])
                sheet_report.cell(row=int(row), column=int(columns) + 1).border = thick_border
                sheet_report.cell(row=int(row), column=int(columns) + 1).font = header_font

        elif columns_titles[columns] == 'fe':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]),
                                                                                   0)  # Writes value to cell
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            # Alarm limits
            # noinspection PyBroadException
            try:
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        fe_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        fe_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        fe_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                    sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')
        elif columns_titles[columns] == 'v':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]),
                                                                                   0)  # Writes value to cell
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            # Alarm limits
            # noinspection PyBroadException
            try:
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        v_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        v_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        v_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                    sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')

        elif columns_titles[columns] == 's':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]) / 60, 2)
            if row == 2:
                sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]) * 100, 2)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            try:
                if row != 2:
                    if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                            s_yellow):
                        sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                    if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                            s_orange):
                        sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                    if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                            s_red):
                        sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')

        elif columns_titles[columns] == 'ni':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]), 0)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            try:
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        ni_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        ni_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        ni_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                    sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')

        elif columns_titles[columns] == 'cr':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]), 0)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            try:
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        cr_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        cr_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        cr_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                    sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border

            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')

        elif columns_titles[columns] == 'pb':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]), 0)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            try:
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        pb_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        pb_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        pb_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                    sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')

        elif columns_titles[columns] == 'cu':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]), 0)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            try:
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        cu_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        cu_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        cu_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                    sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')

        elif columns_titles[columns] == 'ca':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]), 0)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            try:
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        ca_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        ca_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        ca_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                    sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')

        elif columns_titles[columns] == 'zn':
            try:
                if float(vars()[column_data]) > int(
                        zn_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(vars()[column_data]) > int(
                        zn_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(vars()[column_data]) > int(
                        zn_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
            except ValueError:
                print('Value failed for zn')

            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]), 0)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = Border(left=Side(style='thin'),
                                                                                     right=Side(style='medium'),
                                                                                     top=Side(style='medium'),
                                                                                     bottom=Side(style='medium'))
        else:
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]), 0)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
    row += 1
    # sheet_report.cell(row=row_count, column=(column + 1)).value = str(dictionary[id_for_bottle][categories[column]])
os.chdir(excel_folder)
try:
    wb.save('excel_file_name_missing.xlsx')
    file_name = os.path.realpath('excel_file_name_missing.xlsx')
    file_name_new = str(
        'SDA_' + str(latest_date_split[2] + '-' + latest_date_split[1] + '-' + latest_date_split[0]) + '.xlsx')
    try:
        os.unlink(file_name_new)
    except:
        pass

    os.rename(file_name, file_name_new)
except PermissionError:
    print('Was not able to save data to the excel file.\nPlease check that file is closed and try again.')
print('Program has finished. Press any button followed by the Enter button to exit.')
input()
print('Program made by Kenny Kristiansen')

#! python3
# Seamate_converter.py - Converts Seamate data, to excel data

import os
import csv
import openpyxl as excel
import re


seamate_file_location = r'C:\Users\kenny\Desktop'
os.chdir(seamate_file_location)
filenames_in_location = os.listdir(seamate_file_location)

matching_files = []
for i in range(len(filenames_in_location)):
    if str(filenames_in_location[i]).startswith('SEAMATE'):
        matching_files += [filenames_in_location[i]]
        #print('Seamate file found: ' + str(filenames_in_location[i]))


# Find newest file, and use that one for further usage
date_extraction = re.compile(r'(SEAMATE_SAMPLE_TESTS_DATA_)(\d\d)(\w)(\d\d)(\w)(\d\d\d\d)(.*)', re.VERBOSE)

newest_file_index = 0
latest_year = 0
latest_month = 0
latest_date = 0

for i in range(len(matching_files)):
    extracted_data = (date_extraction.search(matching_files[i]))
    year = int(extracted_data.group(6))
    if year >= latest_year:
        latest_year = year
        month = int(extracted_data.group(4))
        if month >= latest_month:
            latest_month = month
            date = int(extracted_data.group(2))
            if date >= latest_date:
                latest_date = date

                newest_file_index = i


file_open = open(matching_files[newest_file_index])
print('Using file with newest date: ' + str(matching_files[newest_file_index]))
print('Date of file: ' + str(latest_date) + '/' + str(latest_month) + '/' + str(latest_year))
file_read = csv.reader(file_open)

print('Converting Seamate file to Excel')
# Write a category sorted dictionary
dictionary = {}
for row in file_read:
    dictionary_values = {}
    if file_read.line_num > 4:
        # Defining categories
        if file_read.line_num == 5:  # Categories
            categories = row
            #print('Categories are: ' + str(row))
            continue

        if file_read.line_num > 6:

            try:
                bottle_id = row[4]
            except:
                # print('No bottle ID found in row #' + str(file_read.line_num))
                continue
            for values in range(len(row)):
                dictionary_values[categories[values]] = str(row[values])


                dictionary[bottle_id] = dictionary_values

dictionary_keys = list(dictionary.keys())

# save to excel sheet
wb = excel.Workbook(False)
sheet_name = wb.sheetnames[0]
sheet = wb.worksheets[0]

for row_1 in range(len(categories)):
    sheet.cell(row=1, column=((row_1)+1)).value = categories[row_1]
print('Bottle entries found: ' + str(len(dictionary_keys)))
for row in range(len(dictionary_keys)):
    # print(dictionary.get(dictionary_keys[row]))
    # print(dictionary_keys[row])
    for column in range(len(categories)):
        id_for_bottle = dictionary_keys[row]
        try:
            sheet.cell(row=(row+2), column=(column+1)).value = str(dictionary[id_for_bottle][categories[column]])
        except:
            #print('Error happened with id: ' + bottle_id)
            continue
wb.save('excel_file_name_missing.xlsx')
print('Seamate file succesfully converted to Excel')
wb.close()


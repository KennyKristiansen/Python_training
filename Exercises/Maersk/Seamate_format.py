#! python3
# Seamate_format.py - Converts seamate excel, into monthly reports
import configparser
import os
import openpyxl as excel
import datetime
from openpyxl import styles
from openpyxl.styles import Border, Side

seamate_file_location = r'C:\Users\kenny\Desktop'
os.chdir(seamate_file_location)
wb = excel.load_workbook('excel_file_name_missing.xlsx', False)
sheet_name = wb.sheetnames[0]

print('Reading settings.txt')
# Read settings file
try:
    settings_file_read = open('settings.txt', mode='r')
except FileNotFoundError:
    print('The settings.txt file was not found.\n\nAn empty settings file has been created.')
    settings_file_write = open('settings.txt', mode='w')
    settings_file_write.write('# Instructions:\n# This file contains all the adjustable limits for the colorization in the excel spreadsheet.\n# If these are left blank, fallback values will be used instead.\n# Please do not remove this file, it is needed for running the program.\n\n[Variables]\nfe_red = 120\nfe_orange = 91\nfe_yellow = 90\nv_red = 800\nv_orange = 700\nv_yellow = 600')

config = configparser.ConfigParser()
config.read("settings.txt")

fe_red = config.get("Variables", "fe_red", fallback=300)
fe_orange = config.get("Variables", "fe_orange", fallback=300)
fe_yellow = config.get("Variables", "fe_yellow", fallback=300)
v_red = config.get("Variables", "v_red", fallback=300)
v_orange = config.get("Variables", "v_orange", fallback=300)
v_yellow = config.get("Variables", "v_yellow", fallback=300)

try:
    wb.sheetnames[1]
except IndexError:
    wb.create_sheet('Monthly', 1)

sheet_data = wb.worksheets[0]
sheet_report = wb.worksheets[1]

print('Identifying latest sample date')
categories = []
for column_count in range(sheet_data.max_column):
    categories.append(str(sheet_data.cell(row=1, column=column_count + 1).value))
group_row = categories.index('GROUP') + 1
date_row = categories.index('SAMPLE DATE') + 1
bottle_row = categories.index('BOTTLE') + 1
location_row = categories.index('LOCATION') + 1
s_row = categories.index('S +/-') + 1
v_row = categories.index('V') + 1
fe_row = categories.index('FE') + 1
ni_row = categories.index('NI') + 1
cr_row = categories.index('CR') + 1
pb_row = categories.index('PB') + 1
cu_row = categories.index('CU') + 1
ca_row = categories.index('CA') + 1
zn_row = categories.index('ZN') + 1

# find latest date

newest_sample = datetime.datetime(1, 1, 1)
latest_sample_date = ()
latest_sample_id = ()
newest_samples = []
latest_date_split = []
for row_count in range(2, sheet_data.max_row):
    date = sheet_data.cell(row=row_count + 1, column=date_row).value
    bottle_id = sheet_data.cell(row=row_count + 1, column=bottle_row).value
    date_split = str.split(date, ' ')

    day = date_split[0]
    if day == 'SAMPLE':
        continue
    month = date_split[1]
    if month == 'Jan':
        month = 1
    if month == 'Feb':
        month = 2
    if month == 'Mar':
        month = 3
    if month == 'Apr':
        month = 4
    if month == 'May':
        month = 5
    if month == 'Jun':
        month = 6
    if month == 'Jul':
        month = 7
    if month == 'Aug':
        month = 8
    if month == 'Sep':
        month = 9
    if month == 'Oct':
        month = 10
    if month == 'Nov':
        month = 11
    if month == 'Dec':
        month = 12

    month = month
    year = date_split[2]
    time = date_split[3]
    # print(day, month, year, time)
    sample_date = datetime.datetime(int(year), int(month), int(day))

    if sample_date > newest_sample:
        newest_samples = []  # Eftersom at nyeste dato er ændret, nulstilles denne liste
        newest_sample = sample_date
        latest_sample_date = sample_date  # Bruges som global værdi for dato
        latest_sample_id = bottle_id  # Bruges som global værdi for ID

    if sample_date == newest_sample:
        newest_samples.append(bottle_id)  # Samler alle de nyeste ID'er i en liste
        latest_date_split = str(str(latest_sample_date).replace(' 00:00:00', '')).split('-')
print('Latest sample date: ' + latest_date_split[2] + '/' + latest_date_split[1] + '/' + latest_date_split[0])

# Excel styles
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))
thick_border = Border(left=Side(style='medium'),
                      right=Side(style='medium'),
                      top=Side(style='medium'),
                      bottom=Side(style='medium'))

red = excel.styles.PatternFill("solid", fgColor="FF0000")
orange = excel.styles.PatternFill("solid", fgColor="FF8000")
yellow = excel.styles.PatternFill("solid", fgColor="FFFF00")
green = excel.styles.PatternFill("solid", fgColor="00FF00")

header_font = excel.styles.Font(bold=True,)
header_alignment = excel.styles.Alignment(horizontal='center')

# Looking up all the wanted values for the newest samples
sheet_report.column_dimensions['A'].width = 12
row = 1
columns_titles = ['location', 'fe', 'v', 's', 'ni', 'cr', 'pb', 'cu', 'ca', 'zn']
for category_count in range(len(columns_titles)):
    sheet_report.cell(row=row, column=category_count + 1).value = columns_titles[category_count]
    sheet_report.cell(row=row, column=category_count + 1).border = thick_border
    sheet_report.cell(row=row, column=category_count + 1).font = header_font
    sheet_report.cell(row=row, column=category_count + 1).alignment = header_alignment

#    sheet_report.cell(row=row, column=(category_count + 1)).value = categories[category_count]

row += 1
for row_count in range(sheet_data.max_row):
    bottle_id = sheet_data.cell(row=row_count + 1, column=bottle_row).value
    bottle_id_correct = False
    for bottle_id_check in range(len(newest_samples)):
        if bottle_id == newest_samples[bottle_id_check]:
            bottle_id_correct = True
    if not bottle_id_correct:
        continue
    location = sheet_data.cell(row=row_count + 1, column=location_row).value
    s = sheet_data.cell(row=row_count + 1, column=s_row).value
    v = sheet_data.cell(row=row_count + 1, column=v_row).value
    fe = sheet_data.cell(row=row_count + 1, column=fe_row).value
    ni = sheet_data.cell(row=row_count + 1, column=ni_row).value
    cr = sheet_data.cell(row=row_count + 1, column=cr_row).value
    pb = sheet_data.cell(row=row_count + 1, column=pb_row).value
    cu = sheet_data.cell(row=row_count + 1, column=cu_row).value
    ca = sheet_data.cell(row=row_count + 1, column=ca_row).value
    zn = sheet_data.cell(row=row_count + 1, column=zn_row).value

    for columns in range(len(columns_titles)):
        column_data = columns_titles[columns]
        # sheet_report.cell(row=int(row), column=int(columns)+1).value = vars()[columns_titles[columns]]
        if columns_titles[columns] == 'location':
            if str(vars()[column_data]) == 'Cylinder Oil Daily Tank':
                sheet_report.cell(row=int(row), column=int(columns) + 1).value = 'Cyl. day tank'
                sheet_report.cell(row=int(row), column=int(columns) + 1).border = thick_border
                sheet_report.cell(row=int(row), column=int(columns) + 1).font = header_font
            else:
                sheet_report.cell(row=int(row), column=int(columns) + 1).value = str(vars()[column_data])
                sheet_report.cell(row=int(row), column=int(columns) + 1).border = thick_border
                sheet_report.cell(row=int(row), column=int(columns) + 1).font = header_font
        elif columns_titles[columns] == 's':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]) / 60, 2)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
        elif columns_titles[columns] == 'fe':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]),
                                                                                   0)  # Writes value to cell
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            # Alarm limits
            # noinspection PyBroadException
            try:
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        fe_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        fe_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        fe_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                    sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')
        elif columns_titles[columns] == 'v':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]),
                                                                                   0)  # Writes value to cell
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            # Alarm limits
            # noinspection PyBroadException
            try:
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        v_yellow):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = yellow
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        v_orange):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = orange
                if float(str(sheet_report.cell(row=int(row), column=int(columns) + 1).value).replace(',', '.')) > int(
                        v_red):
                    sheet_report.cell(row=int(row), column=int(columns) + 1).fill = red
                    sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
            except:
                print('Marking high ' + columns_titles[columns] + ' content failed.')
        elif columns_titles[columns] == 'zn':
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]) / 60, 2)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = Border(left=Side(style='thin'),
                                                                                     right=Side(style='medium'),
                                                                                     top=Side(style='medium'),
                                                                                     bottom=Side(style='medium'))

        else:
            sheet_report.cell(row=int(row), column=int(columns) + 1).value = round(float(vars()[column_data]), 0)
            sheet_report.cell(row=int(row), column=int(columns) + 1).border = thin_border
    row += 1
    # sheet_report.cell(row=row_count, column=(column + 1)).value = str(dictionary[id_for_bottle][categories[column]])
try:
    wb.save('excel_file_name_missing.xlsx')
except PermissionError:
    print('Was not able to save data to the excel file.\nPlease check that file is closed and try again.')
